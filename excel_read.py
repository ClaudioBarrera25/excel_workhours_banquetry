#%%

from openpyxl import Workbook, load_workbook
import numpy as np
import random
import pandas as pd 
import datetime 

def generate_random_time(time_day='Noon'):
    if time_day == 'Noon':
        hour = random.randint(16, 17)
        if hour ==16:
            minute = random.randint(0, 59)
        elif hour == 17:
            minute = random.randint(0, 30)
    if time_day == 'Morning':
        hour = random.randint(7, 8)
        if hour ==7:
            minute = random.randint(0, 59)
        elif hour ==8:
            minute = random.randint(0, 30)
    return f"{hour:02d}:{minute:02d}"



# Función para generar domingos
def generar_domingos(fecha_inicio, fecha_fin):
    """
    Genera una lista de domingos entre dos fechas dadas.
    :param fecha_inicio: str, fecha de inicio en formato 'YYYY-MM-DD'
    :param fecha_fin: str, fecha de fin en formato 'YYYY-MM-DD'
    :return: list, fechas de domingos en formato datetime.date
    """
    # Crear un rango de fechas
    fechas = pd.date_range(start=fecha_inicio, end=fecha_fin, freq='D')
    # Filtrar domingos (weekday = 6)
    domingos = fechas[fechas.weekday == 6]
    # Convertir a lista de fechas en formato date
    return domingos.date.tolist()

# Ejemplo de uso
fecha_inicio = "2023-09-01"
fecha_fin = "2024-12-31"
domingos_mes = generar_domingos(fecha_inicio, fecha_fin)

feriados_chile = [
    # Feriados 2023
    "01/01/2023",  # Año Nuevo
    "07/04/2023",  # Viernes Santo
    "08/04/2023",  # Sábado Santo
    "01/05/2023",  # Día del Trabajador
    "21/05/2023",  # Día de las Glorias Navales
    "26/06/2023",  # San Pedro y San Pablo
    "16/07/2023",  # Día de la Virgen del Carmen
    "15/08/2023",  # Asunción de la Virgen
    "18/09/2023",  # Primera Junta de Gobierno
    "19/09/2023",  # Día de las Glorias del Ejército
    "09/10/2023",  # Encuentro de Dos Mundos (movible)
    "27/10/2023",  # Día de las Iglesias Evangélicas y Protestantes (movible)
    "31/10/2023",  # Día de las Iglesias Evangélicas y Protestantes
    "01/11/2023",  # Día de Todos los Santos
    "08/12/2023",  # Inmaculada Concepción
    "25/12/2023",  # Navidad

    # Feriados 2024
    "01/01/2024",  # Año Nuevo
    "29/03/2024",  # Viernes Santo
    "30/03/2024",  # Sábado Santo
    "01/05/2024",  # Día del Trabajador
    "21/05/2024",  # Día de las Glorias Navales
    "01/07/2024",  # San Pedro y San Pablo
    "16/07/2024",  # Día de la Virgen del Carmen
    "15/08/2024",  # Asunción de la Virgen
    "18/09/2024",  # Primera Junta de Gobierno
    "19/09/2024",  # Día de las Glorias del Ejército
    "14/10/2024",  # Encuentro de Dos Mundos (movible)
    "31/10/2024",  # Día de las Iglesias Evangélicas y Protestantes
    "01/11/2024",  # Día de Todos los Santos
    "08/12/2024",  # Inmaculada Concepción
    "25/12/2024",  # Navidad
]

fechas_feriados_dtime = [datetime.datetime.strptime(date_str, "%d/%m/%Y").date() for date_str in feriados_chile]

fechas_excluir = set(fechas_feriados_dtime + domingos_mes)

#%%


files_subscript = [
    # '2022-08',
    # '2022-09',
    # '2022-10',
    # '2022-11',
    # '2022-12',
    # '2023-01',
    # '2023-02',
    # '2023-03',
    # '2023-04',
    # '2023-05',
    # '2023-06',
    # '2023-07',
    # '2023-08',
    '2023-09',
    '2023-10',
    '2023-11',
    '2023-12',
    '2024-01',
    '2024-02',
    '2024-03',
    '2024-04',
    '2024-05',
    '2024-06',
    '2024-07',
    '2024-08',
    '2024-09',
    '2024-10',
    '2024-11',
    '2024-12'
    ]
subs = -1
current_date_selected = files_subscript[subs]
current_month = datetime.datetime.strptime(current_date_selected, "%Y-%m").month
current_year = datetime.datetime.strptime(current_date_selected, "%Y-%m").year
filepath_to_load = fr'Asistencias 2022 a 2024\xlsx\Todos los informes({current_date_selected}).xlsx'
filepath_to_save = fr'Asistencias 2022 a 2024\completos\Todos los informes({current_date_selected}).xlsx'


workbook = load_workbook(filename=filepath_to_load)
sheetnames = workbook.sheetnames 
registro_asistencia_sheet = workbook['Registro asistencia']
print(filepath_to_load)




#%%




# Check length of the rows 
cell_values_a_row = []
for cell in registro_asistencia_sheet.iter_rows(min_row=1,min_col=1, max_col=1):
    cell_values_a_row.append(cell[0].value)

days_of_month = []
for cell in registro_asistencia_sheet.iter_rows(min_row=4, max_row=4,min_col=1):
    for col in cell:
        if col.value != None:
            days_of_month.append(col.value)

columns_month = []
for cell in registro_asistencia_sheet.iter_rows(min_row=4, max_row=4,min_col=1):
    for col in cell:
        if col.value != None:
            columns_month.append(col.column) 

no_trabaja = ['AlejandraR']
names_employees = {}
for row in registro_asistencia_sheet.iter_rows(min_row=4,min_col=1):
    nombre_str = row[7].value
    nombre_value = row[9].value
    coordenada = row[9].coordinate
    column_letter = row[9].column_letter
    row_name = row[9].row
    row_time = row[9].row +1 
    if nombre_str  == 'Nombre :' and nombre_value in no_trabaja:
        names_employees[row_time] = {'row_name':row_name, 'column':column_letter, 'coordinate': coordenada, 'nombre':nombre_value, 'trabaja': 'No'} 
    if nombre_str  == 'Nombre :' and nombre_value not in no_trabaja:
        names_employees[row_time] = {'row_name':row_name, 'column':column_letter, 'coordinate': coordenada, 'nombre':nombre_value, 'trabaja': 'Si'} 


#%%

print(columns_month)
print(days_of_month)
print(cell_values_a_row)
print(names_employees)


#%%
time_rows = [6, 9, 12, 15, 18, 21, 24, 27, 30, 33, 36, 39, 42, 45, 48, 51, 54]
# Create array to check if all column is none
array_times = []
for row_number in time_rows:
    for row in registro_asistencia_sheet.iter_rows(min_row=row_number, max_row= row_number, max_col=max(columns_month), values_only=True):
        array_times.append(row)
array_times = np.array(array_times)

#%%
# Logic to see if day was worked
# Actualizar `mask_dict` con días no laborales
mask_dict = {}
for column, index_col in zip(array_times.T, columns_month):
    dia = index_col
    dia_actual = datetime.date(current_year, current_month, dia)
    if dia_actual in fechas_excluir:
        mask_dict[index_col] = 'Holiday'
    # if all(value is None for value in column):
    #     mask_dict[index_col] = 'Holiday'  # Considerar vacíos como no trabajado
    else:
        mask_dict[index_col] = 'Work'



#%%

for row_number in time_rows:
    print(names_employees[row_number]['nombre'])
    for row in registro_asistencia_sheet.iter_rows(min_row=row_number, max_row= row_number, max_col=max(columns_month)):
        # each row is one employee
        for cell in row:
            # Check if it is work day 
            if mask_dict[cell.column] == 'Work':
                if cell.value is not None:
                    # Case 1
                    list_times_splitted = cell.value.splitlines()
                    if len(list_times_splitted) == 2:
                        str_to_write = f'{list_times_splitted[0]}' + '\n' + f'{list_times_splitted[-1]}'
                    # Case 2
                    elif len(list_times_splitted) > 2:
                        str_to_write = f'{list_times_splitted[0]}' + '\n' + f'{list_times_splitted[-1]}'
                    # Case 3
                    elif len(list_times_splitted) ==1:
                        str_to_write = f'{generate_random_time("Morning")}' + '\n' + f'{generate_random_time("Noon")}'
                elif cell.value is None:
                    str_to_write = f'{generate_random_time("Morning")}' + '\n' + f'{generate_random_time("Noon")}' 
                # print(str_to_write)
                cell.value = str_to_write
            
            elif mask_dict[cell.column] == 'Holiday':
                continue
            
        # if row[0].value is not None:
        #     print(row[0].value.splitlines())   


#%%

workbook.save(filepath_to_save)


#%%


for current_date_selected in files_subscript:
    current_month = datetime.datetime.strptime(current_date_selected, "%Y-%m").month
    current_year = datetime.datetime.strptime(current_date_selected, "%Y-%m").year
    filepath_to_load = fr'Asistencias 2022 a 2024\xlsx\Todos los informes({current_date_selected}).xlsx'
    filepath_to_save = fr'Asistencias 2022 a 2024\completos\Todos los informes({current_date_selected}).xlsx'


    workbook = load_workbook(filename=filepath_to_load)
    sheetnames = workbook.sheetnames 
    registro_asistencia_sheet = workbook['Registro asistencia']
    print(filepath_to_load)





    # Check length of the rows 
    cell_values_a_row = []
    for cell in registro_asistencia_sheet.iter_rows(min_row=1,min_col=1, max_col=1):
        cell_values_a_row.append(cell[0].value)

    days_of_month = []
    for cell in registro_asistencia_sheet.iter_rows(min_row=4, max_row=4,min_col=1):
        for col in cell:
            if col.value != None:
                days_of_month.append(col.value)

    columns_month = []
    for cell in registro_asistencia_sheet.iter_rows(min_row=4, max_row=4,min_col=1):
        for col in cell:
            if col.value != None:
                columns_month.append(col.column) 

    no_trabaja = ['AlejandraR']
    names_employees = {}
    for row in registro_asistencia_sheet.iter_rows(min_row=4,min_col=1):
        nombre_str = row[7].value
        nombre_value = row[9].value
        coordenada = row[9].coordinate
        column_letter = row[9].column_letter
        row_name = row[9].row
        row_time = row[9].row +1 
        if nombre_str  == 'Nombre :' and nombre_value in no_trabaja:
            names_employees[row_time] = {'row_name':row_name, 'column':column_letter, 'coordinate': coordenada, 'nombre':nombre_value, 'trabaja': 'No'} 
        if nombre_str  == 'Nombre :' and nombre_value not in no_trabaja:
            names_employees[row_time] = {'row_name':row_name, 'column':column_letter, 'coordinate': coordenada, 'nombre':nombre_value, 'trabaja': 'Si'} 



    print(columns_month)
    print(days_of_month)
    print(cell_values_a_row)
    print(names_employees)

    time_rows = [6, 9, 12, 15, 18, 21, 24, 27, 30, 33, 36, 39, 42, 45, 48, 51, 54]
    # Create array to check if all column is none
    array_times = []
    for row_number in time_rows:
        for row in registro_asistencia_sheet.iter_rows(min_row=row_number, max_row= row_number, max_col=max(columns_month), values_only=True):
            array_times.append(row)
    array_times = np.array(array_times)

    # Logic to see if day was worked
    # Actualizar `mask_dict` con días no laborales
    mask_dict = {}
    for column, index_col in zip(array_times.T, columns_month):
        dia = index_col
        dia_actual = datetime.date(current_year, current_month, dia)
        if dia_actual in fechas_excluir:
            mask_dict[index_col] = 'Holiday'
        # if all(value is None for value in column):
        #     mask_dict[index_col] = 'Holiday'  # Considerar vacíos como no trabajado
        else:
            mask_dict[index_col] = 'Work'




    for row_number in time_rows:
        print(names_employees[row_number]['nombre'])
        for row in registro_asistencia_sheet.iter_rows(min_row=row_number, max_row= row_number, max_col=max(columns_month)):
            # each row is one employee
            for cell in row:
                # Check if it is work day 
                if mask_dict[cell.column] == 'Work':
                    if cell.value is not None:
                        # Case 1
                        list_times_splitted = cell.value.splitlines()
                        if len(list_times_splitted) == 2:
                            str_to_write = f'{list_times_splitted[0]}' + '\n' + f'{list_times_splitted[-1]}'
                        # Case 2
                        elif len(list_times_splitted) > 2:
                            str_to_write = f'{list_times_splitted[0]}' + '\n' + f'{list_times_splitted[-1]}'
                        # Case 3
                        elif len(list_times_splitted) ==1:
                            str_to_write = f'{generate_random_time("Morning")}' + '\n' + f'{generate_random_time("Noon")}'
                    elif cell.value is None:
                        str_to_write = f'{generate_random_time("Morning")}' + '\n' + f'{generate_random_time("Noon")}' 
                    # print(str_to_write)
                    cell.value = str_to_write
                
                elif mask_dict[cell.column] == 'Holiday':
                    continue
                
            # if row[0].value is not None:
            #     print(row[0].value.splitlines())   



    workbook.save(filepath_to_save)
# %%
